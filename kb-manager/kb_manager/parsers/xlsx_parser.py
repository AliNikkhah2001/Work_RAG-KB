"""Excel XLSX parser with support for multiple KB schemas.

Two engines are available:
- ``openpyxl`` (default): pure-Python, reliable Unicode (UTF-8) support.
- ``calamine``: Rust-backed reader (python-calamine), fast, full Unicode
  support including Persian/Arabic scripts.

Select with the ``KB_XLSX_ENGINE`` environment variable
(``auto`` | ``openpyxl`` | ``calamine``).  Every parsed string is verified
for Unicode integrity so mojibake (e.g. Persian letters replaced by ``?``)
is surfaced as a warning in metadata instead of silently corrupting the KB.
"""

import os
import re
import logging
from pathlib import Path

from openpyxl import load_workbook

from kb_manager.parsers.base import BaseParser, ParsedDocument

logger = logging.getLogger(__name__)

# Range of Persian/Arabic script codepoints used for integrity checking.
_ARABIC_MIN = 0x0600
_ARABIC_MAX = 0x06FF
_ZWNJ = 0x200C


def _is_arabic_script(ch: str) -> bool:
    """Check if a character belongs to the Persian/Arabic script."""
    return _ARABIC_MIN <= ord(ch) <= _ARABIC_MAX


def _verify_string_integrity(value: str) -> list[str]:
    """Detect real mojibake in a cell string.

    Flags:
    - U+FFFD replacement characters (the reader already saw bad bytes).
    - Control characters other than ``\\n`` / ``\\r`` / ``\\t``.
    - Literal ``?`` sitting between two Arabic-script letters, which is
      the classic signature of text that was decoded with a wrong codepage.

    Deliberately ignores embedded newlines/tabs, which are valid in cells.

    Args:
        value: Cell string to inspect.

    Returns:
        List of human-readable problem descriptions (empty if clean).
    """
    problems: list[str] = []
    if "\ufffd" in value:
        problems.append("contains U+FFFD replacement character")
    for i, ch in enumerate(value):
        code = ord(ch)
        if code < 32 and ch not in "\n\r\t":
            problems.append(f"contains control char U+{code:04X}")
        elif ch == "?" and 0 < i < len(value) - 1:
            prev_arabic = _is_arabic_script(value[i - 1])
            next_arabic = _is_arabic_script(value[i + 1])
            if prev_arabic or next_arabic:
                problems.append("'?' adjacent to Arabic-script letter (mojibake signature)")
    # De-duplicate while preserving order
    return list(dict.fromkeys(problems))


# Known schema column patterns
SCHEMA_A_COLUMNS = {
    "reason_code",
    "model_name",
    "model_id",
    "brief_explanation",
    "detailed_explanation",
    "reason_text",
    "improvement_suggestions",
    "bin_score",
    "bin_impact",
    "feature_score",
    "feature_impact",
    "bin_details",
    "keywords",
    "feature_name",
    "data_source",
}

SCHEMA_B_COLUMNS = {"question", "model", "briefanswer", "answer", "keyword"}

SCHEMA_C_COLUMNS = {
    "documentname",
    "title",
    "sectiontitle",
    "content",
    "type",
    "version",
    "author(s)",
    "heading",
    "keywords",
    "summary",
}


def _normalize_col(name: str) -> str:
    """Normalize column name for comparison."""
    return re.sub(r"[\s_\-]+", "", name.strip().lower())


def _detect_schema(headers: list[str]) -> str | None:
    """Detect which KB schema a sheet uses based on column names.

    Uses a threshold-based match: if >= 60% of a schema's expected columns
    are present, treat it as that schema.  This handles files that are
    missing one optional column (e.g. ``model`` in CRM Q&A).

    Args:
        headers: List of column header strings.

    Returns:
        Schema identifier string or None for generic format.
    """
    normalized = {_normalize_col(h) for h in headers if h}

    schemas = [
        ("reason_codes", SCHEMA_A_COLUMNS),
        ("crm_qa", SCHEMA_B_COLUMNS),
        ("articles", SCHEMA_C_COLUMNS),
    ]
    for name, required in schemas:
        overlap = len(normalized & required)
        if overlap >= len(required) * 0.6:
            return name
    return None


def _is_temp_file(file_path: str) -> bool:
    """Check if file is a temporary Excel lock file."""
    name = Path(file_path).name
    return name.startswith("~$")


class XlsxParser(BaseParser):
    """Parser for Excel XLSX files supporting multiple KB schemas.

    Handles three known schemas:
    - reason_codes: Reason code analysis with model scoring
    - crm_qa: CRM Q&A pairs with keywords
    - articles: Article documentation with sections and metadata

    Falls back to generic key-value parsing for unknown formats.
    """

    def __init__(self, engine: str | None = None) -> None:
        """Initialize the parser.

        Args:
            engine: Engine to use (``openpyxl`` | ``calamine`` | ``auto``).
                Defaults to the ``KB_XLSX_ENGINE`` env var, then ``auto``.
        """
        self.engine = (engine or os.getenv("KB_XLSX_ENGINE", "auto")).lower()
        if self.engine not in ("auto", "openpyxl", "calamine"):
            logger.warning("Unknown KB_XLSX_ENGINE %r, falling back to openpyxl", self.engine)
            self.engine = "openpyxl"

    def can_parse(self, file_path: str) -> bool:
        """Check if the file is an Excel file that can be parsed.

        Args:
            file_path: Path to check.

        Returns:
            True if the file has an xlsx extension and is not a temp file.
        """
        if _is_temp_file(file_path):
            return False
        return Path(file_path).suffix.lower() == ".xlsx"

    def parse(self, file_path: str) -> ParsedDocument:
        """Parse an Excel file and extract content from all sheets.

        Args:
            file_path: Path to the XLSX file.

        Returns:
            ParsedDocument with sheets data and combined text content.

        Raises:
            FileNotFoundError: If the file does not exist.
            ValueError: If the file cannot be opened or parsed.
        """
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        if _is_temp_file(file_path):
            raise ValueError(f"Cannot parse temporary Excel file: {file_path}")

        abs_path = str(Path(file_path).resolve())
        title = Path(file_path).stem

        sheets_data: list[dict] = []
        all_content: list[str] = []
        integrity_issues: list[str] = []
        engine_used = self.engine

        try:
            sheet_names, rows_by_sheet = self._read_sheets(abs_path)
            if engine_used == "auto":
                engine_used = self._detect_available_engine()
            for sheet_name in sheet_names:
                sheet_result = self._parse_sheet_rows(
                    sheet_name, rows_by_sheet[sheet_name], integrity_issues
                )
                if sheet_result is not None:
                    sheets_data.append(sheet_result)
                    all_content.append(self._sheet_to_text(sheet_result))
        except ValueError:
            raise
        except Exception as e:
            raise ValueError(f"Failed to open Excel file with engine '{self.engine}': {e}") from e

        if not sheets_data:
            raise ValueError(f"No valid sheets found in {file_path}")

        combined_content = "\n\n".join(all_content)
        metadata: dict = {
            "sheet_count": len(sheets_data),
            "parser_engine": engine_used,
        }
        if integrity_issues:
            metadata["integrity_warnings"] = integrity_issues
            logger.warning(
                "Unicode integrity issues in %s: %d cell(s) - %s",
                Path(file_path).name,
                len(integrity_issues),
                integrity_issues[0],
            )

        return ParsedDocument(
            source_path=abs_path,
            title=title,
            content=combined_content,
            file_type="xlsx",
            metadata=metadata,
            sheets=sheets_data,
            sections=None,
        )

    def _detect_available_engine(self) -> str:
        """Return the best engine installed on this machine."""
        try:
            import python_calamine  # noqa: F401

            return "calamine"
        except ImportError:
            return "openpyxl"

    def _read_sheets(self, abs_path: str) -> tuple[list[str], dict[str, list[list[object]]]]:
        """Read all sheets using the configured engine.

        Returns:
            (sheet_names, rows_by_sheet) where rows are lists of raw values.
        """
        if self.engine == "auto":
            if self._detect_available_engine() == "calamine":
                return self._read_sheets_calamine(abs_path)
            return self._read_sheets_openpyxl(abs_path)
        if self.engine == "calamine":
            return self._read_sheets_calamine(abs_path)
        return self._read_sheets_openpyxl(abs_path)

    def _read_sheets_calamine(self, abs_path: str) -> tuple[list[str], dict[str, list[list[object]]]]:
        """Read sheets with the Rust-backed python-calamine reader."""
        from python_calamine import CalamineWorkbook

        wb = CalamineWorkbook.from_path(abs_path)
        names = wb.sheet_names
        out: dict[str, list[list[object]]] = {}
        for name in names:
            ws = wb.get_sheet_by_name(name)
            out[name] = ws.to_python()
        return names, out

    def _read_sheets_openpyxl(self, abs_path: str) -> tuple[list[str], dict[str, list[list[object]]]]:
        """Read sheets with openpyxl (read-only streaming)."""
        wb = load_workbook(abs_path, read_only=True, data_only=True)
        try:
            out: dict[str, list[list[object]]] = {}
            for name in wb.sheetnames:
                ws = wb[name]
                out[name] = [list(row) for row in ws.iter_rows(values_only=True)]
            return wb.sheetnames, out
        finally:
            wb.close()

    def _parse_sheet_rows(
        self,
        sheet_name: str,
        rows: list[list[object]],
        integrity_issues: list[str],
    ) -> dict | None:
        """Parse rows of a single worksheet and detect its schema.

        Args:
            sheet_name: Name of the worksheet.
            rows: Raw rows of cell values.
            integrity_issues: Accumulator for Unicode-integrity warnings.

        Returns:
            Dictionary with sheet data or None if sheet is empty/invalid.
        """
        if not rows:
            return None

        header_row = rows[0]
        if header_row is None:
            return None

        headers = [str(h).strip() if h is not None else "" for h in header_row]

        # Filter out empty header columns at the end
        while headers and not headers[-1]:
            headers.pop()

        if len(headers) < 2:
            return None

        # Collect remaining rows
        data_rows: list[list[str]] = []
        for row in rows[1:]:
            if row is None:
                continue
            values: list[str] = []
            for i, cell in enumerate(row):
                if i >= len(headers):
                    break
                values.append(self._format_cell(cell))
            # Skip fully empty rows
            if any(v for v in values):
                data_rows.append(values)
                for i, v in enumerate(values):
                    if not v:
                        continue
                    problems = _verify_string_integrity(v)
                    if problems:
                        col_label = headers[i] if i < len(headers) else str(i + 1)
                        integrity_issues.append(
                            f"sheet={sheet_name!r} column={col_label!r}: {problems[0]}"
                        )

        if not data_rows:
            return None

        schema = _detect_schema(headers)

        return {
            "name": sheet_name,
            "headers": headers,
            "rows": data_rows,
            "schema": schema,
        }

    @staticmethod
    def _format_cell(cell: object) -> str:
        """Normalize a single cell value to string."""
        if cell is None:
            return ""
        if isinstance(cell, float):
            return f"{cell:g}" if cell == int(cell) else str(cell)
        return str(cell).strip()

    def _sheet_to_text(self, sheet_data: dict) -> str:
        """Convert sheet data to readable text format.

        Multi-line cell values are collapsed to a single line so that
        the pipe-delimited row format stays intact for the chunker.

        Args:
            sheet_data: Sheet dictionary with name, headers, rows, and schema.

        Returns:
            Formatted text representation of the sheet.
        """
        lines: list[str] = [f"=== Sheet: {sheet_data['name']} ==="]

        if sheet_data["schema"]:
            lines.append(f"[Schema: {sheet_data['schema']}]")
            lines.append("")

        headers = sheet_data["headers"]
        rows = sheet_data["rows"]

        for row in rows:
            parts: list[str] = []
            for header, value in zip(headers, row):
                if value and value.strip():
                    # Collapse multi-line cells into a single line
                    collapsed = value.strip().replace("\n", " ").replace("\r", "")
                    parts.append(f"{header}: {collapsed}")
            if parts:
                lines.append(" | ".join(parts))

        return "\n".join(lines)
