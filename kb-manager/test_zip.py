import zipfile
import os

with zipfile.ZipFile(r"C:\Users\10225\Downloads\KB\31Tir1405(1).zip", 'r') as z:
    for info in z.infolist():
        if 'EtebaritoProblems' in info.filename:
            with open("zip_info.txt", "w", encoding="utf-8") as f:
                f.write(f"Filename: {info.filename}\n")
                f.write(f"Size: {info.file_size}\n")
            z.extract(info, "test_extract")
            break

# Now test the extracted file
import openpyxl
extracted_path = os.path.join("test_extract", info.filename)
wb = openpyxl.load_workbook(extracted_path, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))

with open("zip_test.txt", "w", encoding="utf-8") as f:
    f.write(f"Headers: {list(rows[0])}\n")
    for i, r in enumerate(rows[1:2]):
        f.write(f"---ROW {i+1}---\n")
        for h, v in zip(rows[0], r):
            if v:
                f.write(f"  {h}: {repr(str(v))[:200]}\n")

wb.close()
print("done")