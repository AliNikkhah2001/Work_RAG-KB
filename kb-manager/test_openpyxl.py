from openpyxl import load_workbook

wb = load_workbook(r"C:\Users\10225\Downloads\KB\extracted\اشکالات سامانه اعتباریتو\EtebaritoProblems.xlsx", read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))

with open("openpyxl_test.txt", "w", encoding="utf-8") as f:
    f.write(f"Headers: {list(rows[0])}\n")
    for i, r in enumerate(rows[1:3]):
        f.write(f"---ROW {i+1}---\n")
        for h, v in zip(rows[0], r):
            if v:
                f.write(f"  {h}: {repr(str(v))[:200]}\n")

wb.close()