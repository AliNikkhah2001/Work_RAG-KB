"""Map IVA pairs to expected chunk ids = ALL chunks in the document holding the golden answer."""
import asyncio
import json
import os
import sys
import unicodedata

os.environ["KB_DB_URL"] = "sqlite+aiosqlite:///D:/Code/KB/kb-manager/data/kb_1405.db"
sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import load_workbook
from sqlalchemy import text

from kb_manager.config import load_config
from kb_manager.models.database import Database


def norm(s: str) -> str:
    s = unicodedata.normalize("NFC", s)
    s = s.replace("\u200c", " ").replace("\u064a", "\u06cc").replace("\u0643", "\u06a9")
    for fa, en in zip("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789"):
        s = s.replace(fa, en)
    return "".join(c for c in s if c not in "؟?،;,.!").strip().lower()


def toks(s: str) -> set[str]:
    return set(norm(s).split())


def main():
    pairs = []
    wb = load_workbook(
        r"D:/Code/KB/kb-source/1405-05-31/TestQuestions_IVA/InitialTestQuestion.xlsx",
        read_only=True, data_only=True,
    )
    ws = wb["Sheet1"]
    for row in ws.iter_rows(values_only=True):
        q, a = row[0], row[1]
        if q and str(q).strip() and a and str(a).strip():
            pairs.append((str(q).strip(), str(a).strip()))
    wb.close()

    async def _load():
        cfg = load_config()
        db = Database(cfg.db)
        rows = []
        async with db.session() as s:
            r = await s.execute(
                text("SELECT id, document_id, content, chunk_type FROM chunks")
            )
            for row in r.fetchall():
                rows.append(
                    {"id": row[0], "doc": row[1], "content": row[2], "ct": toks(row[2]), "type": row[3]}
                )
        await db.close()
        return rows

    rows = asyncio.run(_load())
    by_doc: dict[str, list[dict]] = {}
    for ch in rows:
        by_doc.setdefault(ch["doc"], []).append(ch)

    out = []
    # For duplicate handling (Q15): group chunks by content hash, so any doc with same answer is valid
    from collections import defaultdict
    content_to_docs: dict[frozenset, set[str]] = defaultdict(set)
    for ch in rows:
        content_to_docs[frozenset(ch["ct"])].add(ch["doc"])

    for i, (q, a) in enumerate(pairs):
        at = toks(a)
        # find document(s) with max average answer coverage
        best_doc = None
        best_doc_cov = -1.0
        for doc, chunks in by_doc.items():
            best_chunk_cov = max(
                (len(at & c["ct"]) / max(len(at), 1)) if c["ct"] and at else 0.0 for c in chunks
            )
            if best_chunk_cov > best_doc_cov:
                best_doc_cov, best_doc = best_chunk_cov, doc
        # Include ALL docs that share the same content hash as best_doc's best chunk (fixes Q15 duplicate miss)
        # Find the content hash of the best chunk in best_doc
        best_chunk_ct = max((c["ct"] for c in by_doc[best_doc]), key=lambda ct: len(at & ct) / max(len(at),1))
        # All docs that have a chunk with identical token set are considered valid
        valid_docs = {doc for doc, chunks in by_doc.items() if any(c["ct"] == best_chunk_ct for c in chunks)}
        # Also include docs with same content hash (for exact duplicates like "جزئیات قراردادهای منفی...")
        # Fallback to single doc if no duplicate
        if not valid_docs:
            valid_docs = {best_doc}
        doc_ids = []
        for doc in valid_docs:
            doc_ids.extend(c["id"] for c in by_doc[doc])
        out.append(
            {
                "query": q,
                "expected_chunk_ids": doc_ids,
                "expected_answer": a,
                "format": "verbatim",
                "category": "factual",
                "ans_cov": round(best_doc_cov, 2),
                "expected_docs": len(valid_docs),
            }
        )
        print(i + 1, "doc_cov", round(best_doc_cov, 2), "n_id", len(doc_ids), "valid_docs", len(valid_docs))

    dest = r"D:/Code/KB/kb-manager/data/test_questions_iva.json"
    with open(dest, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print("saved", dest)


if __name__ == "__main__":
    main()