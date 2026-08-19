import zipfile
import openpyxl
import io

zip_path = r"C:\Users\10225\Downloads\KB\31Tir1405.zip"

with zipfile.ZipFile(zip_path, "r", allowZip64=True) as zf:
    for name in zf.namelist():
        if "IndividualAndBusinessCreditReport" in name:
            with open(r"C:\Users\10225\Downloads\KB\test_report_result.txt", "w", encoding="utf-8") as out:
                out.write(f"Testing: {name}\n")
                with zf.open(name) as f:
                    content = f.read()
                    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                    ws = wb[wb.sheetnames[0]]
                    rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
                    out.write(f"Headers: {list(rows[0]) if rows else 'EMPTY'}\n")
                    for i, r in enumerate(rows[1:3]):
                        out.write(f"Row {i+1}: {r}\n")
                    wb.close()
            break

print("done")