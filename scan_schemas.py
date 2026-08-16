import openpyxl
import os
import json

path = r'C:\Users\10225\Downloads\KB\extracted'
results = {}

for root, dirs, files in os.walk(path):
    for f in files:
        if f.endswith('.xlsx') and not f.startswith('~$'):
            full = os.path.join(root, f)
            rel = os.path.relpath(full, path)
            try:
                wb = openpyxl.load_workbook(full, read_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
                results[rel] = {
                    "sheets": wb.sheetnames if hasattr(wb, 'sheetnames') else [],
                    "columns": list(rows[0]) if rows else [],
                    "row_count": len(rows) - 1 if len(rows) > 1 else 0,
                    "sample_row": list(rows[1]) if len(rows) > 1 else [],
                }
            except Exception as e:
                results[rel] = {"error": str(e)}

with open(r'C:\Users\10225\Downloads\KB\file_schemas.json', 'w', encoding='utf-8') as f:
    json.dump(results, f, ensure_ascii=False, indent=2, default=str)

print(f"Processed {len(results)} files")
